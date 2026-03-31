from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import warnings
warnings.filterwarnings('ignore')
import openpyxl
import sys

#configuração do arquivo excel

options = Options()
options.add_argument("--start-maximized")
options.add_argument('log-level=3')#para ignorar warnings
options.add_experimental_option('excludeSwitches', ['enable-logging'])#para ignorar warnings
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 120)
wait_faster = WebDriverWait(driver, 2)
wait_fast = WebDriverWait(driver, 0.5)



book = openpyxl.load_workbook('Dados.xlsx')
consultation_page = book['Consulta']

def ler_credenciais(arquivo):
    with open(arquivo, 'r', encoding='utf-8') as f:
        linhas = f.read().splitlines()
        return linhas[0], linhas[1]  # Retorna login e senha

# lendo credenciais do arquivo
credenciais_arquivo = 'credenciais.txt'
v8_login, v8_senha = ler_credenciais(credenciais_arquivo)


#abrindo a v8 e realizando login
url = "https://app.v8sistema.com/fgts"
driver.get(url)
field = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[1]/div/button'))).click()
time.sleep(0.20)

field = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="1-email"]')))
field.clear()  # limpa o campo antes de enviar o texto
field.send_keys(v8_login)

field = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="1-password"]'))).send_keys(v8_senha)
time.sleep(0.20)

field = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="1-submit"]'))).click()
time.sleep(0.20)

loading = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/div[3]/div/div[2]/div[1]/p[1]')))
#time.sleep(3)


time.sleep(0.20)
button_fgts = driver.find_element(By.XPATH, '//*[text()="FGTS"]').click() #botao do FGTS
time.sleep(0.20)
arrow_options = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div[3]/main/div/form/div/div[2]/div[1]/div/div[1]/div/div[1]/div[1]/div/div'))).click()
time.sleep(0.20)
options = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="field-:rg:"] | //*[@id="react-select-2-listbox"]')))
time.sleep(0.20)
field = wait.until(EC.presence_of_element_located((By.XPATH, f"//div[text()='BMS']"))).click()
time.sleep(0.20)


#fazendo a consulta
#instituição
def config_consulta(driver, wait, wait_fast):
    options_tabs = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tabs-:rb:--tabpanel-0"]/div/div[1]/div/div[1]/div[2]/div/div/div[2]'))).click()
    time.sleep(0.20)
    field = wait_fast.until(EC.presence_of_element_located((By.XPATH, f"//div[text()='Cometa']"))).click()
    time.sleep(0.20)

    #seguro
    options_safe = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="tabs-:rb:--tabpanel-0"]/div/div[1]/div/div[1]/div[3]/div/div/div[2]'))).click()
    time.sleep(0.20)
    field = wait_fast.until(EC.presence_of_element_located((By.XPATH, f"//div[text()='Não']"))).click()
    time.sleep(0.20)

#chamando a função
config_consulta(driver, wait, wait_fast)

#função de consulta
def consultar_cpf(campo_cpf):

    campo_cpf = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div[4]/div[1]/div[1]/div[2]/div/input')))
    time.sleep(0.20)
    campo_cpf.send_keys(Keys.CONTROL + "a")  # seleciona todo o conteúdo
    time.sleep(0.20)    
    campo_cpf.send_keys(Keys.BACKSPACE)  # apaga o conteúdo
    time.sleep(0.20)
    campo_cpf.send_keys(cpf_consulta)
    time.sleep(0.20)
    campo_cpf = wait_fast.until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div[4]/div[1]/div[1]/div[2]/button'))).click() #limpar cache
    time.sleep(1)
    campo_cpf = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="field-:rf:"]')))
    time.sleep(0.20)
    campo_cpf.send_keys(Keys.CONTROL + "a")  
    time.sleep(0.20)
    campo_cpf.send_keys(Keys.BACKSPACE)  
    time.sleep(0.20)
    campo_cpf.send_keys(cpf_consulta)
    time.sleep(0.20)
    time.sleep(1)
    campo_cpf.send_keys(Keys.RETURN)
    time.sleep(0.20)
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element_located((By.XPATH, '//*[@id="tabs-:rf:--tabpanel-0"]/div/div[1]/button/div | //*[@id="tabs-:r2:--tabpanel-0"]/div/div[1]/button/div/div'))
    )
    print(f"CPF {cpf_consulta} inserido com sucesso")

    time.sleep(2)


def resultados_consulta(driver, wait, row, consultation_page):
    
    if driver.find_elements(By.XPATH, '//*[@id="chakra-toast-manager-top-right"]/div/div'):

        cpf_excel = consultation_page.cell(row=row, column=1).value
        notificação = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="chakra-toast-manager-top-right"]/div/div/div | //*[@id="chakra-toast-manager-top-right"]/div/div/div')))
        texto_notificação = notificação.text
        print(texto_notificação)
        if 'informado não possui contas de' in texto_notificação:
            texto_notificação = 'Sem FGTS'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        elif 'Erro ao consultar saldo, Trabalhador não possui adesão ao saque aniversário vigente na data corrente.' in texto_notificação:
            texto_notificação = 'Sem Adesao'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        elif f'{cpf_consulta} | Não foi possível consultar o saldo no momento! - Instituição Fiduciária não possui autorização do Trabalhador para Operação Fiduciária.' in texto_notificação:
            texto_notificação = 'Não Autorizado'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        elif 'Saldo insuficiente, parcelas menores R$100,00' in texto_notificação:
            texto_notificação = 'Sem saldo'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        elif 'Erro ao consultar saldo, Não foi possível consultar o saldo no momento! - Mudanças cadastrais na conta do FGTS foram realizadas, que impedem a contratação. Entre em contato com o setor de FGTS da CAIXA.' in texto_notificação:
            texto_notificação = 'Mudanças Cadastrais'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        elif f'{cpf_consulta} | Não foi possível consultar o saldo no momento! - Operação não permitida antes de 03/02/2026.' in texto_notificação:
            texto_notificação = 'Aniversariante'
            resultado = texto_notificação
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado
        
        elif 'Erro ao consultar saldo, Limite de requisições excedido, tente novamente mais tarde' in texto_notificação or 'Erro ao consultar saldo, Não foi possível consultar o saldo no momento!' or 'Erro ao consultar saldo, undefined' or 'Tente novamente' in texto_notificação:
            campo_cpf = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="field-:rf:"]')))
            time.sleep(10)
            campo_cpf.send_keys(Keys.RETURN)
            time.sleep(3)
            resultados_consulta(driver, wait, row, consultation_page)

        
        

    if not driver.find_elements(By.XPATH, '//*[@id="chakra-toast-manager-top-right"]/div/div/div'):
        print(f'Notificação não encontrada')
        time.sleep(2)
        try:
            cpf_v8 = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div[4]/div[2]/div[1]/table/tbody[1]/tr/td[1]')))
            cpf_excel = consultation_page.cell(row=row, column=1).value
            cpf_excel = cpf_excel.replace('.', '').replace('-', '')

            texto_cpfv8 = cpf_v8.text
            texto_cpfv8 = texto_cpfv8.replace('.', '').replace('-', '')

            print(f'Cpf do excel: {cpf_excel}')
            print(f'CPF da v8:{texto_cpfv8}')

            if cpf_excel == texto_cpfv8:
                texto_cpfv8 = cpf_v8.text
                print(texto_cpfv8)
                resultado = "Com saldo"
                consultation_page.cell(row=row, column=2, value=resultado)

                saldo = wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="main"]/div/div[4]/div[2]/div[1]/table/tbody[1]/tr/td[4]/span')))
                saldo_valor = saldo.text
                print(f'Saldo Encontrado: {saldo_valor}')
                consultation_page.cell(row=row, column=3, value=saldo_valor)
                time.sleep(1)
                config_consulta(driver, wait, wait_fast)
                return resultado

        except Exception as e2:
            # caso o elemento não seja encontrado ou não tenha sido atualizado
            print(f'Erro ao verificar a atualização do elemento: {e2}')
            resultado = 'Erro Desconhecido'
            consultation_page.cell(row=row, column=2, value=resultado)
            return resultado




for row in range(2, consultation_page.max_row + 1):
    cpf_consulta = consultation_page.cell(row=row, column=1).value
    cpf_consulta = str(cpf_consulta).zfill(11).replace('.', '').replace('-', '')
    
    if not cpf_consulta:
        print(f"⚠️ CPF inválido ou vazio na linha {row}. Encerrando o programa.")
        consultation_page.cell(row=row, column=2, value="CPF Inválido")  # coloca cpf invalido na segunda coluna
        book.save('Dados.xlsx')
        sys.exit()  # saindo do programa

    consultar_cpf(cpf_consulta)
    resultado = resultados_consulta(driver, wait, row, consultation_page)
    book.save('Dados.xlsx')
    time.sleep(5)

book.save('Dados.xlsx')
    